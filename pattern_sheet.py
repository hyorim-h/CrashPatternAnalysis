import pandas as pd
from pattern_divide import *
from aggregate_input import *
from openpyxl import load_workbook
import os
import warnings
warnings.filterwarnings(action='ignore')

cpath = os.path.dirname((os.path.abspath('__file__')))
dtf = pd.read_csv(cpath + '/산출물/교통사고패턴매칭.csv', encoding='cp949')
os.mkdir(cpath+'/산출물/사고패턴시트')

# 패턴화 항목들
party1 = {"AM":"승용차", "TR":"화물차", "BU":"승합차", "TW":"이륜차", "BC":"자전거","MC":"원동기장치자전거","CM":"건설기계","PM":"개인형이동수단(PM)","SP":"특수차","AT":"사륜오토바이(ATV)","FM":"농기계"}
crash_types = {"C":"횡단중", 'R':"차도통행중", 'S':"길가장자리구역통행중", 'P':'보도통행중'}
road_types = {"E":'고속국도', "G":"군도", "H":"일반국도", 'M':"특별광역시도", 'R':"지방도", 'S':"시도"}
road_shapes = {'I1':"교차로내", 'I2':"교차로횡단보도내", 'I3':"교차로부근", 'R1':"기타단일로", 'R2':"지하차도(도로)내", 'R3':"교량위", 'R4':"고가도로위", 'R5':"터널안"}
behavior1 ={'D1':"직진 중", 'D2':"진로변경 중", 'D3':"앞지르기 중", 'P1':"주정차중", 'R1':"후진 중", 'S1':"주행 중 대기", 'T1':"좌회전 중", 'T2':"우회전 중", 'T3':"U턴 중"}
behavior2 = {'C1':"횡단보도내", 'C2':"횡단보도외", 'O1':"승차 중 관련", 'O2':"하차 중 관련", 'P1':"보도통행", 'R1':"등지고 통행", 'R2':"마주보고 통행", 'R3':"도로위 작업 중", 'R4':"도로위 놀이 중", 'S1':"길가장자리구역 통행"}

for pt in party1.keys():
    for ct in crash_types.keys():
        npath = cpath + '/산출물/사고패턴시트/' + pt + 'PE-' + ct
        os.mkdir(npath)
        for rt in road_types.keys():
            for rs in road_shapes.keys():
                # 패턴시트 집계를 위한 sample file 열기
                wb = load_workbook(cpath + '/패턴시트sample.xlsx')
                ws = wb.active

                for b1 in behavior1.keys():
                    for b2 in behavior2.keys():
                        pattern = pt + "PE" + '-' + ct + '-'+ rt + rs + '-' + b1 + b2
                        temp = dtf[(dtf['party']==pt+"PE")&(dtf['crash']==ct)&(dtf['road']==rt+rs)&(dtf['behavior']==b1+b2)]
                        if len(temp)!=0:
                            sht_name = b1 + b2
                            sht = wb[wb.sheetnames[0]]
                            ws = wb.copy_worksheet(sht)
                            ws.title = sht_name
                            print(sht_name)

                            # 사고 개요 input
                            input_crash_summary(ws, temp, pattern, party1[pt], crash_types[ct], road_types[rt], road_shapes[rs], behavior1[b1], behavior2[b2])

                            # 집계 결과 input
                            # 사고내용
                            crash_result = aggregate_result(temp)
                            input_aggregate_result(ws, crash_result)

                            # 사망자수
                            fat_result = fat_aggregate_result(temp)
                            input_fat_aggregate_result(ws, fat_result)

                            # 주야 input
                            daynight = daynight_aggregate_result(temp)
                            input_daynight_aggregate_result(ws, daynight)

                            # 시간대 input
                            time_area = time_aggregate_result(temp)
                            input_time_aggregate_result(ws, time_area)

                            # 기상상태 input
                            weather = weather_aggregate_result(temp)
                            input_weather_aggregate_result(ws, weather)

                            # 노면상태 input
                            road_cond = roadcond_aggregate_result(temp)
                            input_roadcond_aggregate_result(ws, road_cond)

                            # 운전자법규위반 input
                            law = law_aggregate_result(temp)
                            input_law_aggregate_result(ws, law)
                            
                            # 도로선형 input
                            roadshape2 = roadshape2_aggregate_result(temp)
                            input_roadshape2_aggregate_result(ws, roadshape2)
                            
                            # 교차로형태 input
                            intersection = intersection_aggregate_result(temp)
                            input_intersection_aggregate_result(ws, intersection)
                            
                            #중앙분리시설 input
                            roadsep = roadsep_aggegate_result(temp)
                            input_roadsep_aggregate_result(ws, roadsep)
                if len(wb.sheetnames) > 1:
                    ws = wb[wb.sheetnames[0]]
                    wb.remove(ws)

                    output_filename = '/' + pt + "PE" + '-' + ct + '-'+ rt + rs + '.xlsx'
                    wb.save(npath + output_filename)      
        
        if len(os.listdir(npath))==0:
            os.rmdir(npath)
